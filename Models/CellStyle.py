import constants
from Helpers import JsonHelper
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, Alignment, Protection


class AlignmentCustom:
    def __init__(self, horizontal=None, vertical=None, text_rotation=None, wrap_text=None, shrink_to_fit=None,
                 indent=None):
        self.indent = indent
        self.shrink_to_fit = shrink_to_fit
        self.wrap_text = wrap_text
        self.text_rotation = text_rotation
        self.vertical = vertical
        self.horizontal = horizontal


class FillCustom:
    def __init__(self, fill_type=None, start_color=None, end_color=None):
        self.end_color = end_color
        self.start_color = start_color
        self.fill_type = fill_type


class FontCustom:
    def __init__(self, name=None, bold=None, size=None, italic=None, vert_align=None, underline=None, strike=None,
                 color=None):
        self.color = color
        self.name = name
        self.underline = underline
        self.vert_align = vert_align
        self.italic = italic
        self.size = size
        self.bold = bold
        self.strike = strike


class BorderCustom:
    def __init__(self, border_style=None, color=None):
        self.color = color
        self.border_style = border_style


class ProtectionCustom:
    def __init__(self, locked=None, hidden=None):
        self.hidden = hidden
        self.locked = locked


class CellStyleCustom:
    def __init__(self, name=None, font=None, alignment=None, border=None, fill=None, protection=None):
        self.protection = protection
        self.name = name
        self.font = font
        self.alignment = alignment
        self.border = border
        self.fill = fill


def read_styles():
    styles = JsonHelper.read_json(constants.STYLE_JSON_PATH)
    cell_style_list = []

    for style in styles:
        name = style[constants.STYLE_NAME]
        font = FontCustom(**style[constants.STYLE_FONT]) if constants.STYLE_FONT in style else None
        border = BorderCustom(**style[constants.STYLE_BORDER]) if constants.STYLE_BORDER in style else None
        fill = FillCustom(**style[constants.STYLE_FILL]) if constants.STYLE_FILL in style else None
        alignment = AlignmentCustom(**style[constants.STYLE_ALIGNMENT]) if constants.STYLE_ALIGNMENT in style else None
        protection = ProtectionCustom(
            **style[constants.STYLE_PROTECTION]) if constants.STYLE_PROTECTION in style else None
        cell_style = CellStyleCustom(name=name, font=font, alignment=alignment, border=border, fill=fill,
                                     protection=protection)
        cell_style_list.append(cell_style)
    return cell_style_list


def register_style(style, book):
    style_name = style.name
    style_font = style.font
    style_fill = style.fill
    style_alignment = style.alignment
    style_border = style.border
    style_protection = style.protection

    wb = book.book
    if style_name in wb.named_styles:
        del wb._named_styles[wb.style_names.index(style_name)]

    registered_style = NamedStyle(name=style_name)

    if style_font is not None:
        registered_style.font = Font(name=style_font.name, size=style_font.size, bold=style_font.bold,
                                     italic=style_font.italic, vertAlign=style_font.vert_align,
                                     underline=style_font.underline,
                                     strike=style_font.strike, color=style_font.color)

    if style_border is not None:
        bd = Side(style=style_border.border_style, color=style_border.color)
        registered_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    if style_fill is not None:
        registered_style.fill = PatternFill(fill_type=style_fill.fill_type, start_color=style_fill.start_color,
                                            end_color=style_fill.end_color)

    if style_alignment is not None:
        registered_style.alignment = Alignment(horizontal=style_alignment.horizontal, vertical=style_alignment.vertical,
                                               text_rotation=style_alignment.text_rotation,
                                               wrap_text=style_alignment.wrap_text,
                                               shrink_to_fit=style_alignment.shrink_to_fit,
                                               indent=style_alignment.indent)

    if style_protection is not None:
        registered_style.protection = Protection(locked=style_protection.locked, hidden=style_protection.hidden)

    return registered_style


def register_styles(book):
    style_list = read_styles()
    for style in style_list:
        named_style = register_style(style, book)
        book.book.add_named_style(named_style)


def create_pattern_fill(start_color, end_color, fill_type):
    return PatternFill(start_color=start_color, end_color=end_color, fill_type=fill_type)
