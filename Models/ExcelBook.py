from openpyxl import load_workbook, Workbook

class ExcelBook:
    def __init__(self, path):
        self.path = path
        try:
            self.book = load_workbook(path)
            print("Opening existing Excel file: ", path)
        except:
            self.book = Workbook()
            self.book.save(path)
            print("Creating new Excel file: ", path)


