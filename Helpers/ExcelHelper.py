import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def write_df_to_excel(df, excel_book, sheet_name, start_row, start_col):
    book = excel_book.book
    path = excel_book.path
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        try:
            writer.book.remove(writer.book[sheet_name])
            print("Worksheet: ", sheet_name, " already existed")
            print("Deleting old sheet to save new one ...")
        except:
            print("Saving new worksheet: ", sheet_name)
        finally:
            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, startcol=start_col, engine='openpyxl')


def adjust_column_width(excel_book, sheet_name):
    path = excel_book.path
    book = excel_book.book
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        # Auto-adjust columns' width
        sheet = book[sheet_name]
        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length * 1.23


def insert_rows(excel_book, sheet_name, insert_row_index, num_of_rows):
    path = excel_book.path
    book = excel_book.book
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        ws = excel_book.book[sheet_name]
        while num_of_rows > 0:
            ws.insert_rows(insert_row_index)
            num_of_rows = num_of_rows - 1


def insert_rows_at_end(excel_book, sheet_name, num_of_rows):
    path = excel_book.path
    book = excel_book.book
    ws = excel_book.book[sheet_name]
    insert_row_index = ws.max_row + 1
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        while num_of_rows > 0:
            ws.insert_rows(insert_row_index)
            num_of_rows = num_of_rows - 1


def write_to_cell(excel_book, sheet_name, cell_row, cell_column, value, cell_style):
    path = excel_book.path
    book = excel_book.book
    ws = excel_book.book[sheet_name]
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        cell = ws.cell(row=cell_row, column=cell_column, value=value)
        if cell_style:
            cell.style = cell_style


def get_letter_of_column(index):
    return get_column_letter(index)


def comparing_conditional_formatting(excel_book, sheet_name, compared_cell, comparing_cell, less_than_comparing_cell_fill, greater_than_comparing_cell_fill):
    ws = excel_book.book[sheet_name]
    ws.conditional_formatting.add(compared_cell, CellIsRule(operator='greaterThan', formula=[comparing_cell], fill=greater_than_comparing_cell_fill))
    ws.conditional_formatting.add(compared_cell, CellIsRule(operator='lessThan', formula=[comparing_cell], fill=less_than_comparing_cell_fill))


# this function returns the formula in the cell (including prefix equal sign), not the real value
def read_cell_formula(excel_book, sheet_name, cell_row, cell_column):
    ws = excel_book.book[sheet_name]
    return ws.cell(cell_row, cell_column).value


# this function returns the formula in the cell (excluding prefix equal sign), not the real value, for the purpose to be nested in another formula
def read_cell_formula_without_equal_sign(excel_book, sheet_name, cell_row, cell_column):
    cell_formula = read_cell_formula(excel_book, sheet_name, cell_row, cell_column)
    return cell_formula[1:]
